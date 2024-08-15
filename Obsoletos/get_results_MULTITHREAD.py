#!
# -*- coding: cp1252 -*-

import os
from typing import Dict, Any, Union

from openpyxl import load_workbook
from datetime import datetime
from statistics import mean
import numpy as np
from time import sleep
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import reduce
from get_variable_name import get_variable_name

from statsmodels.tsa.ar_model import AutoReg
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.vector_ar.var_model import VAR
from statsmodels.tsa.statespace.varmax import VARMAX
from statsmodels.tsa.holtwinters import SimpleExpSmoothing
from statsmodels.tsa.holtwinters import ExponentialSmoothing

from functools import wraps
import warnings

warnings.filterwarnings('ignore', 'statsmodels.tsa.arima_model.ARMA',
                        FutureWarning)
warnings.filterwarnings('ignore', 'statsmodels.tsa.arima_model.ARIMA',
                        FutureWarning)


def measure(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        t = datetime.now()
        result = func(*args, **kwargs)
        print(func.__name__, 'tardó ', str(datetime.now() - t))
        return result

    return wrapper


@measure
def autorregresion(*args):
    """
    Autorregresión
    :param args: lista con la serie de datos a predecir
    :return:
    """
    combinacion = []
    for lista in args:
        # fit model
        model = AutoReg(lista, lags=1, old_names=False)
        model_fit = model.fit()

        # make prediction
        yhat = model_fit.predict(len(lista), len(lista))
        num_selec = round(yhat[0])
        combinacion.append(num_selec)
    return combinacion


@measure
def moving_average(*args):
    """
    Media móvil
    :param args:
    :return:
    """
    combinacion = []
    for data in args:
        # fit model
        model = ARIMA(data, order=(1, 1, 2))
        model_fit = model.fit()
        # make prediction
        yhat = model_fit.predict(len(data), len(data))
        num_selec = round(yhat[0])
        combinacion.append(num_selec)
    return combinacion


@measure
def autorregresion_ma(*args):
    """
    Media móvil autorregresiva
    :param args:
    :return:
    """
    combinacion = []
    for data in args:
        # fit model
        model = SARIMAX(data, order=(1, 0, 0))
        model_fit = model.fit(disp=False)
        # make prediction
        yhat = model_fit.predict(len(data), len(data))
        num_selec = round(yhat[0])
        combinacion.append(num_selec)
    return combinacion


@measure
def season_autore_ma(*args):
    """
    Média móvil integrada autorregresiva
    :param args:
    :return:
    """
    combinacion = []
    for data in args:
        # fit model
        model = SARIMAX(data, order=(1, 1, 1), seasonal_order=(0, 0, 0, 0))
        model_fit = model.fit(disp=False)
        # make prediction
        yhat = model_fit.predict(len(data), len(data))
        num_selec = round(yhat[0])
        combinacion.append(num_selec)

    return combinacion


@measure
def season_autore_ma_ex_reg(data1, data2, exogeno):
    """
    Média móvil integrada autorregresiva estacional con regresores exógenos
    :return:
    """
    # Ajusto las series de datos a listas con el mismo número de elementos
    len1 = len(data1)
    len2 = len(data2)
    if len1 > len2:
        data1 = data1[len(data1):len(data1) - len(data2) - 1:-1]
    elif len2 > len1:
        data2 = data2[len(data2):len(data2) - len(data1) - 1:-1]

    combinacion = []
    # fit model
    model = SARIMAX(data1, exog=data2, order=(1, 1, 1), seasonal_order=(0, 0, 0, 0))
    model_fit = model.fit(disp=False)
    # make prediction
    yhat = model_fit.predict(len(data1), len(data1), exog=exogeno)
    num_selec = round(yhat[0])
    combinacion.append(num_selec)

    return combinacion


@measure
def ar_vectorial(data1, data2):
    """
    Autorregresión vectorial
    :return:
    """
    data1 = np.array(data1, dtype=np.float64)
    data2 = np.array(data2, dtype=np.float64)
    data = list(zip(data2[::-1], data1[::-1]))
    combinacion = []
    # fit model
    model = VAR(data)
    model_fit = model.fit()
    # make prediction
    yhat = model_fit.forecast(model_fit.endog, steps=1)
    num_selec = round(mean(*yhat.tolist()))
    combinacion.append(num_selec)

    return combinacion


@measure
def ar_vectorial_mm(data1, data2):
    """
    Media móvil con Autorregresión vectorial
    :return:
    """
    data1 = np.array(data1, dtype=np.float64)
    data2 = np.array(data2, dtype=np.float64)
    data = list(zip(data2[::-1], data1[::-1]))
    data = [list(x) for x in data]
    combinacion = []
    # fit model
    model = VARMAX(data, order=(1, 1))
    model_fit = model.fit(disp=False)
    # make prediction
    yhat = model_fit.forecast()
    num_selec = round(mean(*yhat.tolist()))
    combinacion.append(num_selec)

    return combinacion


@measure
def ar_vectorial_mm_er(data1, data2, data3):
    """
    Media móvil con Autorregresión vectorial y regresores exógenos
    :param data3: lista con tantos datos como data1 y data2
    :return:
    """
    dataavg = mean(data3)
    data1 = np.array(data1, dtype=np.float64)
    data2 = np.array(data2, dtype=np.float64)
    data = list(zip(data2[::-1], data1[::-1]))
    exogeno = [x for x in data3[::-1]]
    exogeno = exogeno[:min(len(data1), len(data2))]
    combinacion = []
    # fit model
    model = VARMAX(data, exog=exogeno, order=(1, 1))
    model_fit = model.fit(disp=False)
    # make prediction
    data_exog2 = [[dataavg]]
    yhat = model_fit.forecast(exog=data_exog2)
    num_selec = round(mean(*yhat.tolist()))
    combinacion.append(num_selec)

    return combinacion


@measure
def atenuacion_exponencial(*args, formato=0):
    """
    Atenuación exponencial simple
    :param : formato: Si vale 0 devuelve un entero entre 1 y 50. Si no, se entiende que es un porcentaje
    y se multiplica por 100
    :return:
    """
    combinacion = []
    for data in args:
        data = np.array(data, dtype=np.float64)
        # fit model
        model = SimpleExpSmoothing(data)
        model_fit = model.fit()
        # make prediction
        yhat = model_fit.predict(len(data), len(data))
        # print('ATENUACION EXPONENCIAL SIMPLE', yhat, type(yhat))
        # num_selec = yhat
        num_selec = round(np.mean(*yhat.tolist())) if formato == 0 else np.mean(*yhat.tolist()) * 100
        # num_selec = round(mean(*yhat)) if formato == 0 else mean(*yhat)*100
        combinacion.append(num_selec)

    return combinacion


@measure
def atenuacion_exponencial_hw(*args, formato=0):
    """
    Atenuación exponencial de Holt Winter
    :param:
    :return:
    """
    combinacion = []
    for data in args:
        data = np.array(data, dtype=np.float64)
        # fit model
        model = ExponentialSmoothing(data)
        model_fit = model.fit()
        # make prediction
        yhat = model_fit.predict(len(data), len(data))
        num_selec = round(np.mean(*yhat.tolist())) if formato == 0 else np.mean(*yhat.tolist()) * 100
        combinacion.append(num_selec)

    return combinacion


def estadisticas(*milista):
    # resultados = []
    for lista in milista:
        sorteos = len(lista)
        print('sorteos realizados: ', sorteos)
        if sorteos == 0:
            return
        numeros = list(set(lista))

        def extrae_frec(elemento):
            return elemento[1]

        frecuencia = []
        frec_acum = 0
        for numero in numeros:
            veces = lista.count(numero)
            frec = round(veces / sorteos * 100, 3)
            stats = (numero, frec)
            frec_acum += frec
            frecuencia.append(stats)
        avgfreq = round(frec_acum / len(numeros), 3)
        print('Frecuencia media:', avgfreq)
        frecuencia.sort(key=extrae_frec, reverse=True)
        freq_sobre_media = []
        min_elem = 6
        for elem in frecuencia:
            q_elem = len(freq_sobre_media)
            if elem[1] >= avgfreq or q_elem < min_elem:
                freq_sobre_media.append(elem)
        # freq_sobre_media = list(filter(lambda x: x[1] >= avgfreq, frecuencia))
        freq_sobre_media.sort(key=extrae_frec)
        print(freq_sobre_media)
    return


def imprimeparimpar(titulo: str, data2print: list, qdata: int = 10):
    print('\n\t', titulo.upper())
    print('\t', '================================')
    for data in data2print[0:qdata]:
        print(data[0], '  --------->  ', data[1])
    return


def calcula_combinaciones():
    filename = r'primitiva.xlsx'
    fileadr = r'C:/Users/josemanuel.santiago/OneDrive - Uponor Corporation/Documents/JSC/Loterias/Primitiva/'

    primifullfilename = os.path.abspath(fileadr + filename)
    wb = load_workbook(filename=primifullfilename, read_only=True, data_only=True)
    # print(type(wb.sheetnames), wb.sheetnames)
    wb.active = 0
    allcombs = wb.active

    def parimpar(*args):
        evenodd = []
        for arg in args:
            if arg:
                if arg % 2 == 0:
                    evenodd.append('e')
                else:
                    evenodd.append('o')
            else:
                evenodd.append(None)
        return evenodd

    combinaciones = {}
    # Los resultados empiezan a partir de la Fila 3 (row=3)
    for fila in allcombs.iter_rows(min_row=3):
        if type(fila[1].value) is not int:
            break
        if type(fila[0].value) is datetime and 1 <= fila[1].value <= 49:
            combinaciones[fila[0].value.strftime('%A %d-%m-%Y')] = \
                {
                    'n1': fila[1].value,
                    'n2': fila[2].value,
                    'n3': fila[3].value,
                    'n4': fila[4].value,
                    'n5': fila[5].value,
                    'n6': fila[6].value,
                    'cm': fila[7].value,
                    're': fila[8].value,
                    'fn1t': fila[11].value,
                    'fn2t': fila[12].value,
                    'fn3t': fila[13].value,
                    'fn4t': fila[14].value,
                    'fn5t': fila[15].value,
                    'fn6t': fila[16].value,
                    'fcmt': fila[17].value,
                    'fret': fila[18].value,
                    'fn1j': fila[19].value,
                    'fn2j': fila[20].value,
                    'fn3j': fila[21].value,
                    'fn4j': fila[22].value,
                    'fn5j': fila[23].value,
                    'fn6j': fila[24].value,
                    'fcmj': fila[25].value,
                    'frej': fila[26].value,
                    'fn1s': fila[27].value,
                    'fn2s': fila[28].value,
                    'fn3s': fila[29].value,
                    'fn4s': fila[30].value,
                    'fn5s': fila[31].value,
                    'fn6s': fila[32].value,
                    'fcms': fila[33].value,
                    'fres': fila[34].value,
                    'parimpar': parimpar(fila[1].value, fila[2].value, fila[3].value, fila[4].value,
                                         fila[5].value, fila[6].value, fila[7].value, fila[8].value)
                }
    n1list = []
    n1jueves = []
    n1sabado = []
    n2list = []
    n2jueves = []
    n2sabado = []
    n3list = []
    n3jueves = []
    n3sabado = []
    n4list = []
    n4jueves = []
    n4sabado = []
    n5list = []
    n5jueves = []
    n5sabado = []
    n6list = []
    n6jueves = []
    n6sabado = []
    cmlist = []
    cmjueves = []
    cmsabado = []
    relist = []
    rejueves = []
    resabado = []
    fn1t = []
    fn2t = []
    fn3t = []
    fn4t = []
    fn5t = []
    fn6t = []
    fcmt = []
    fret = []
    fn1j = []
    fn2j = []
    fn3j = []
    fn4j = []
    fn5j = []
    fn6j = []
    fcmj = []
    frej = []
    fn1s = []
    fn2s = []
    fn3s = []
    fn4s = []
    fn5s = []
    fn6s = []
    fcms = []
    fres = []
    parimparlist = []
    parimparlistj = []
    parimparlists = []

    for clave, valor in enumerate(combinaciones):
        # print(clave, valor, combinaciones[valor])
        n1list.append(combinaciones[valor]['n1'])
        n2list.append(combinaciones[valor]['n2'])
        n3list.append(combinaciones[valor]['n3'])
        n4list.append(combinaciones[valor]['n4'])
        n5list.append(combinaciones[valor]['n5'])
        n6list.append(combinaciones[valor]['n6'])
        cmlist.append(combinaciones[valor]['cm'])
        if combinaciones[valor]['re']:
            relist.append(combinaciones[valor]['re'])
        fn1t.append(combinaciones[valor]['fn1t'])
        fn2t.append(combinaciones[valor]['fn2t'])
        fn3t.append(combinaciones[valor]['fn3t'])
        fn4t.append(combinaciones[valor]['fn4t'])
        fn5t.append(combinaciones[valor]['fn5t'])
        fn6t.append(combinaciones[valor]['fn6t'])
        fcmt.append(combinaciones[valor]['fcmt'])
        fret.append(combinaciones[valor]['fret'])
        parimparlist.append(combinaciones[valor]['parimpar'])
        if 'thursday' in valor.lower():
            n1jueves.append(combinaciones[valor]['n1'])
            n2jueves.append(combinaciones[valor]['n2'])
            n3jueves.append(combinaciones[valor]['n3'])
            n4jueves.append(combinaciones[valor]['n4'])
            n5jueves.append(combinaciones[valor]['n5'])
            n6jueves.append(combinaciones[valor]['n6'])
            cmjueves.append(combinaciones[valor]['cm'])
            if combinaciones[valor]['re']:
                rejueves.append(combinaciones[valor]['re'])
            fn1j.append(combinaciones[valor]['fn1j'])
            fn2j.append(combinaciones[valor]['fn2j'])
            fn3j.append(combinaciones[valor]['fn3j'])
            fn4j.append(combinaciones[valor]['fn4j'])
            fn5j.append(combinaciones[valor]['fn5j'])
            fn6j.append(combinaciones[valor]['fn6j'])
            fcmj.append(combinaciones[valor]['fcmj'])
            frej.append(combinaciones[valor]['frej'])
            parimparlistj.append(combinaciones[valor]['parimpar'])
        else:
            n1sabado.append(combinaciones[valor]['n1'])
            n2sabado.append(combinaciones[valor]['n2'])
            n3sabado.append(combinaciones[valor]['n3'])
            n4sabado.append(combinaciones[valor]['n4'])
            n5sabado.append(combinaciones[valor]['n5'])
            n6sabado.append(combinaciones[valor]['n6'])
            cmsabado.append(combinaciones[valor]['cm'])
            if combinaciones[valor]['re']:
                resabado.append(combinaciones[valor]['re'])
            fn1s.append(combinaciones[valor]['fn1s'])
            fn2s.append(combinaciones[valor]['fn2s'])
            fn3s.append(combinaciones[valor]['fn3s'])
            fn4s.append(combinaciones[valor]['fn4s'])
            fn5s.append(combinaciones[valor]['fn5s'])
            fn6s.append(combinaciones[valor]['fn6s'])
            fcms.append(combinaciones[valor]['fcms'])
            fres.append(combinaciones[valor]['fres'])
            parimparlists.append(combinaciones[valor]['parimpar'])

    print('-------------------\tNº TOTAL DE SORTEOS ANALIZADO:\t', len(n1list), '\t-------------------')

    c_totales = (n1list, n2list, n3list, n4list, n5list, n6list, cmlist, relist)
    c_jueves = (n1jueves, n2jueves, n3jueves, n4jueves, n5jueves, n6jueves, cmjueves, rejueves)
    c_sabado = (n1sabado, n2sabado, n3sabado, n4sabado, n5sabado, n6sabado, cmsabado, resabado)
    f_totales = (fn1t, fn2t, fn3t, fn4t, fn5t, fn6t, fcmt, fret)
    f_jueves = (fn1j, fn2j, fn3j, fn4j, fn5j, fn6j, fcmj, frej)
    f_sabado = (fn1s, fn2s, fn3s, fn4s, fn5s, fn6s, fcms, fres)
    setparimpartot = list(set(tuple(comb) for comb in parimparlist))
    setparimparjue = list(set(tuple(comb) for comb in parimparlistj))
    setparimparsab = list(set(tuple(comb) for comb in parimparlists))

    parimpartotales: dict[Union[tuple, Any], int] = {}
    for comb in setparimpartot:
        parimpartotales[comb] = parimparlist.count(list(comb))
    ord_totales = sorted(parimpartotales.items(), key=lambda x: x[1], reverse=True)

    parimparjueves = {}
    for comb in setparimparjue:
        parimparjueves[comb] = parimparlistj.count(list(comb))
    ord_jueves = sorted(parimparjueves.items(), key=lambda x: x[1], reverse=True)

    parimparsabado = {}
    for comb in setparimparsab:
        parimparsabado[comb] = parimparlists.count(list(comb))
    ord_sabado = sorted(parimparsabado.items(), key=lambda x: x[1], reverse=True)

    print('Estadísticas TOTALES')
    imprimeparimpar('combinaciones par-impar totales', ord_totales)
    estadisticas(*c_totales)

    print('Estadísticas jueves')
    imprimeparimpar('combinaciones par-impar jueves', ord_jueves)
    estadisticas(*c_jueves)

    print('Estadísticas sabado')
    imprimeparimpar('combinaciones par-impar sábado', ord_sabado)
    estadisticas(*c_sabado)

    artot = ('Autorregresión TOTALES', autorregresion(*c_totales))
    arjueves = ('Autorregresión jueves', autorregresion(*c_jueves))
    arsabado = ('Autorregresión sabado', autorregresion(*c_sabado))

    mov_med_tot = ('Promedio móvil TOTALES', moving_average(*c_totales))
    mov_med_jueves = ('Promedio móvil jueves', moving_average(*c_jueves))
    mov_med_sabado = ('Promedio móvil sabado', moving_average(*c_sabado))

    mov_ar_tot = ('Promedio móvil autorregresivo TOTALES', autorregresion_ma(*c_totales))
    mov_ar_jueves = ('Promedio móvil autorregresivo jueves', autorregresion_ma(*c_jueves))
    mov_ar_sabado = ('Promedio móvil autorregresivo sabado', autorregresion_ma(*c_sabado))

    est_mov_ar_tot = ('Estacional autorregresivo con Promedio móvil integrado TOTALES',
                      season_autore_ma(*c_totales))
    est_mov_ar_jueves = ('Estacional autorregresivo con Promedio móvil integrado jueves',
                         season_autore_ma(*c_jueves))
    est_mov_ar_sabado = ('Estacional autorregresivo con Promedio móvil integrado sabado',
                         season_autore_ma(*c_sabado))

    print(artot, '\n', arjueves, '\n', arsabado, '\n', mov_med_tot, '\n', mov_med_jueves, '\n', mov_med_sabado, '\n',
          mov_ar_tot, '\n', mov_ar_jueves, '\n', mov_ar_sabado, '\n', est_mov_ar_tot, '\n', est_mov_ar_jueves, '\n',
          est_mov_ar_sabado)

    # comb_sarimax = []
    # n1_exog = [max(n1list)]
    # n2_exog = [max(n2list)]
    # n3_exog = [max(n3list)]
    # n4_exog = [max(n4list)]
    # n5_exog = [max(n5list)]
    # n6_exog = [max(n6list)]
    # cm_exog = [max(cmlist)]
    # re_exog = [max(relist)]
    # comb_sarimax.append(*season_autore_ma_ex_reg(n1sabado, n1jueves, n1_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(n2sabado, n2jueves, n2_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(n3sabado, n3jueves, n3_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(n4sabado, n4jueves, n4_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(n5sabado, n5jueves, n5_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(n6sabado, n6jueves, n6_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(cmsabado, cmjueves, cm_exog))
    # comb_sarimax.append(*season_autore_ma_ex_reg(resabado, rejueves, re_exog))
    #
    # comb_arvect = [*ar_vectorial(n1jueves, n1sabado), *ar_vectorial(n2jueves, n2sabado),
    #                *ar_vectorial(n3jueves, n3sabado), *ar_vectorial(n4jueves, n4sabado),
    #                *ar_vectorial(n5jueves, n5sabado), *ar_vectorial(n6jueves, n6sabado),
    #                *ar_vectorial(cmjueves, cmsabado), *ar_vectorial(rejueves, resabado)]
    #
    # comb_arvect_mm = [*ar_vectorial_mm(n1jueves, n1sabado), *ar_vectorial_mm(n2jueves, n2sabado),
    #                   *ar_vectorial_mm(n3jueves, n3sabado), *ar_vectorial_mm(n4jueves, n4sabado),
    #                   *ar_vectorial_mm(n5jueves, n5sabado), *ar_vectorial_mm(n6jueves, n6sabado),
    #                   *ar_vectorial_mm(cmjueves, cmsabado), ar_vectorial_mm(rejueves, resabado)]
    #
    # comb_arvect_mm_er = [*ar_vectorial_mm_er(n1jueves, n1sabado, n1list),
    #                      *ar_vectorial_mm_er(n2jueves, n2sabado, n2list),
    #                      *ar_vectorial_mm_er(n3jueves, n3sabado, n3list),
    #                      *ar_vectorial_mm_er(n4jueves, n4sabado, n4list),
    #                      *ar_vectorial_mm_er(n5jueves, n5sabado, n5list),
    #                      *ar_vectorial_mm_er(n6jueves, n6sabado, n6list),
    #                      *ar_vectorial_mm_er(cmjueves, cmsabado, cmlist),
    #                      *ar_vectorial_mm_er(rejueves, resabado, relist),]
    #
    # print('Media móvil integrada autorregresiva estacional con regresores exógenos,', comb_sarimax)
    # print('Autorregresión vectorial,', comb_arvect)
    # print('Media Móvil con Autorregresión vectorial,', comb_arvect_mm)
    # print('Media Móvil con Autorregresión vectorial y regresores exógenos,', comb_arvect_mm_er)

    sestot = ('Atenuación Exponencial Simple TOTALES,', atenuacion_exponencial(*c_totales))
    sesjueves = ('Atenuación Exponencial Simple jueves,', atenuacion_exponencial(*c_jueves))
    sessabado = ('Atenuación Exponencial Simple sabado,', atenuacion_exponencial(*c_sabado))

    seshwtot = ('Atenuación Exponencial Holt Winter TOTALES,', atenuacion_exponencial_hw(*c_totales))
    seshwjueves = ('Atenuación Exponencial Holt Winter jueves,', atenuacion_exponencial_hw(*c_jueves))
    seshwsabado = ('Atenuación Exponencial Holt Winter sabado,', atenuacion_exponencial_hw(*c_sabado))

    print(sestot, '\n', sesjueves, '\n', sessabado, '\n', seshwtot, '\n', seshwjueves, '\n', seshwsabado, '\n')

    sestot_f = ('Atenuación Exponencial Simple TOTALES', atenuacion_exponencial(*f_totales, formato=1))
    sesjueves_f = ('Atenuación Exponencial Simple jueves', atenuacion_exponencial(*f_jueves, formato=1))
    sessabado_f = ('Atenuación Exponencial Simple sabado', atenuacion_exponencial(*f_sabado, formato=1))

    seshwtot_f = ('Atenuación Exponencial Holt Winter TOTALES', atenuacion_exponencial_hw(*f_totales, formato=1))
    seshwjueves_f = ('Atenuación Exponencial Holt Winter jueves', atenuacion_exponencial_hw(*f_jueves, formato=1))
    seshwsabado_f = ('Atenuación Exponencial Holt Winter sabado', atenuacion_exponencial_hw(*f_sabado, formato=1))

    print(sestot_f, '\n', sesjueves_f, '\n', sessabado_f, '\n', seshwtot_f, '\n', seshwjueves_f, '\n',
          seshwsabado_f, '\n')


def main():
    calcula_combinaciones()


if __name__ == '__main__':
    main()
