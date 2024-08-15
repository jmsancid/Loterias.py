#!
# -*- coding: cp1252 -*-

from openpyxl import load_workbook
from datetime import datetime
import os


def add_combination(combi):
    """
    Añade una combinación al fichero de primitiva. El formato es un diccionario
    con la fecha como clave y las combinaciones como valor
    :param combi: diccionario con fecha de combinación como clave y lista con la combinación como valor
    :return: True si añadida
            False no añadida
    """
    filename = r'primitiva.xlsx'
    fileadr = r'C:/Users/josemanuel.santiago/OneDrive - Uponor Corporation/Documents/JSC/Loterias/Primitiva/'

    primifullfilename = os.path.abspath(fileadr + filename)
    print('... cargando primitiva.xlsx')
    # wb = load_workbook(filename=primifullfilename, read_only=False, data_only=True)
    wb = load_workbook(filename=primifullfilename, read_only=False, data_only=True)
    print('... fichero cargado')
    # print(type(wb.sheetnames), wb.sheetnames)
    wb.active = 0
    allcombs = wb.active
    # Recorro la primera columna. Si la fecha es posterior a la de la fila leída y en la fila leída hay
    # una combinación, sigo bajando. Si la fecha del argumento se encuentra entre 2 fechas existentes,
    # se inserta una fila con la combinación.
    #
    q_combinaciones = len(combi)
    fichero_modificado = False
    print('... se han encontrado ', q_combinaciones, ' combinaciones a comprobar')
    # for comb_n in range(q_combinaciones):
    for comb_n in reversed(combi):
        # dia, mes, anno = list(combi.keys())[comb_n].split('/')
        dia, mes, anno = comb_n.split('/')
        # n1, n2, n3, n4, n5, e1, e2 = list(combi.values())[comb_n]
        n1, n2, n3, n4, n5, n6, compl, reint = combi[comb_n]
        fecha2find = datetime(int(anno), int(mes), int(dia))
        print('... combinación de web a verificar', fecha2find, n1, n2, n3, n4, n5, compl, reint)

        # Buscamos última fila con datos en el excel
        ultimafilacondatos = 3
        for fila in allcombs.iter_rows(min_row=3, min_col=1, max_col=1):
            if not fila[0].value:
                break
            else:
                ultimafilacondatos = fila[0].row
        print('n de filas actualizado: ', ultimafilacondatos)

        # Buscamos hacia atrás la primera fecha anterior a la última leída en la web
        # print('Buscando combinación con fecha anterior a ', fecha2find)
        for fila in reversed(range(ultimafilacondatos + 1)):  # Sumo 1 porque range no tomaría la última fila
            if fila == 3:
                break
            # print('Fila en For: ', fila)
            fechaleida = allcombs.cell(fila, 1).value
            # print('... comparando con fecha ', fechaleida)
            if fechaleida > fecha2find:
                continue
            elif fechaleida == fecha2find:
                break
            elif fechaleida < fecha2find:
                if allcombs.cell(fila + 1, 2).value:  # La fila siguiente es de fecha posterior y tiene combinación
                    # ESTA PARTE DEL CÓDIGO FALLA, CREO QUE POR PROBLEMAS DE MEMORIA
                    print('Insertar combinación en fila ', fila + 1)
                    allcombs.insert_rows(fila + 1)  # insert_rows funciona con el índice de fila, que empieza en 0
                else:
                    print('Añadir combinación en fila ', fila + 1)
                fichero_modificado = True
                allcombs.cell(fila + 1, 1).value = fecha2find
                allcombs.cell(fila + 1, 2).value = n1
                allcombs.cell(fila + 1, 3).value = n2
                allcombs.cell(fila + 1, 4).value = n3
                allcombs.cell(fila + 1, 5).value = n4
                allcombs.cell(fila + 1, 6).value = n5
                allcombs.cell(fila + 1, 7).value = n6
                allcombs.cell(fila + 1, 8).value = compl
                allcombs.cell(fila + 1, 9).value = reint
                break  # No hay que seguir revisando filas
    if fichero_modificado:
        print('... guardando primitiva.xlsx')
        wb.save(primifullfilename)
        print('... cerrando primitiva.xlsx tras escribir')
    else:
        print('... cerrando primitiva.xlsx sin modificar')

    wb.close()
    return fichero_modificado

